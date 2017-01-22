from django.shortcuts import render, redirect, render_to_response, get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect, HttpResponseBadRequest, JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.template import Context, RequestContext
from django.template.loader import get_template
from itertools import chain
from .models import *
from . import serializers
from imagetrac_docker.taskmanager.models import *
from .forms import *
from django.forms.models import modelformset_factory, BaseModelFormSet, inlineformset_factory
from django.template.response import TemplateResponse
from django.views.generic import CreateView, UpdateView
from django.db.models import Q, Count
from itertools import chain
from django.contrib import messages
from django.template.defaultfilters import slugify
import os, glob, shutil, sys, os.path, datetime, time, smtplib, csv
from datetime import timedelta, time
from rest_framework import generics
import django_excel as excel
import pyexcel


# class ListProducts(APIView):
# 	def get(self, request, format=None):
# 		products = RumbaProduct.objects.all()[:100]
# 		serializer = serializers.ProductSerializer(products, many=True)
# 		return Response(serializer.data)

class ListCreateProducts(generics.ListCreateAPIView):
	queryset = RumbaProduct.objects.all()[:100]
	serializer_class = serializers.ProductSerializer

class RetrieveUpdateDestroyProducts(generics.RetrieveUpdateDestroyAPIView):
	queryset = RumbaProduct.objects.all()
	serializer_class = serializers.ProductSerializer

class ListCreateHotItem(generics.ListCreateAPIView):
	queryset = HotItem.objects.all()
	serializer_class = serializers.ItemSerializer

class RetrieveUpdateDestroyHotItem(generics.RetrieveUpdateDestroyAPIView):
	queryset = HotItem.objects.all()
	serializer_class = serializers.ItemSerializer


date = datetime.date.today()

def getCurrMonth(date):
		date = str(date)
		y = str(date[2:4])
		m = str(date[5:7])
		d = str(date[-2:])
		return m,y

datelist = getCurrMonth(date)
cmonth = datelist[0]
if cmonth == '12':
	nmonth = '01'
else:
	nmonth = str(int(cmonth) + 1)
if len(nmonth) == 1:
	nmonth = "0" + nmonth

if cmonth == '01':
	lmonth = '12'
else:
	lmonth = str(int(cmonth) - 1)
if len(lmonth) == 1:
	lmonth = "0" + lmonth

if nmonth == '12':
	fmonth = '01'
else:
	fmonth = str(int(nmonth) + 1)
if len(fmonth) == 1:
	fmonth = "0" + fmonth

cyear = datelist[1]


def login_user(request):
	username = password = ''
	if request.POST:
		username = request.POST['username']
		password = request.POST['password']

		user = authenticate(username=username, password=password)
		if user is not None:
			if user.is_active:
				login(request, user)
				SignUpTask.delay(6, 7)
				return HttpResponseRedirect('/main/')
	return render(request, 'registration/login.html')

def logout_page(request):
	logout(request)
	return redirect('registration/login.html')

##################

# def handler404(request):
#     response = render_to_response('404.html', {},
#                                   context_instance=RequestContext(request))
#     response.status_code = 404
#     return response


# def handler500(request):
#     response = render_to_response('500.html', {},
#                                   context_instance=RequestContext(request))
#     response.status_code = 500
#     return response

def page_not_found(request):
	response = render(request, '404.html')

	response.status_code = 404

	return response

class UploadFileForm(forms.Form):
	file = forms.FileField()


def record_deployed(request):
	if request.method == "POST":
		user = request.user
		files = request.POST.get('textfield', None)
		files = files.split(' ')
		file_dict = {}
		for file_ in files:
			item_no = file_[:19].replace('_', ' ')
			file_dict[item_no] = file_


		output_file = open('/Users/kevin/b5crontab/temp/deployed_record.csv', 'w')
		writer = csv.writer(output_file)

		for item_no, file_ in file_dict.items():
			writer.writerow([item_no, file_, user])

		output_file.close()
		return HttpResponseRedirect('/main/')  
	else:
		return render(request, 'upload_forms/prompt.html')

def upload(request):
	if request.method == "POST":
		form = UploadFileForm(request.POST, request.FILES)
		if form.is_valid():
			filehandle = request.FILES['file']
			return excel.make_response(filehandle.get_sheet(), "csv")
	else:
		form = UploadFileForm()
		context = {'form': form}
	return render(request, 'dex/dex_upload.html', context)


def import_data(request):
	if request.method == "POST":
		form = UploadFileForm(request.POST,
							  request.FILES)
		
		if form.is_valid():
			request.FILES['file'].save_to_database(
				model=DexProduct,
				mapdict=['ad_date', 'item_ns', 'vendor_number', "item_no", "mfg", "desc", "notes", "order_date", "received_dc", "received_137", "received_buyer", "received_other", "photo_dldate", "whowhen", "studio_out", "checked_out", "confirmed_placed", "shooting_instructions", "studio_in"])
			messages.success(request, 'records successfully imported to database.')
			return HttpResponseRedirect('/dexedit/')
		else:
			return HttpResponseBadRequest()

	else:
		form = UploadFileForm()
		context = {'form': form, 'title': 'Import excel data into database example', 'header': 'Please upload sample-data.xls:'}
	return render(request, 'upload_forms/dex_upload.html', context)


def import_first(request):
	if request.method == "POST":
		form = UploadFileForm(request.POST,
							  request.FILES)
		
		if form.is_valid():
			request.FILES['file'].save_to_database(
				model=FirstReceipt,
				mapdict=["buyer", "dc_received_u", "curr_dc_oh_u", "dc_curr_oo_u", "date_received", "item_ns", "short_sku", "item_no", "mfg", "description", "have_image", "ad_date", "order_date", "received_dc", "received_137", "from_file", "photo_dldate", "whowhen", "studio_out", "checked_out", "confirmed_placed", "studio_in", "merch_to_137"])
			messages.success(request, 'records successfully imported to database.')
			return HttpResponseRedirect('/main/')
		else:
			return HttpResponseBadRequest()
	else:
		form = UploadFileForm()
		context = {'form': form, 'title': 'Import excel data into database example', 'header': 'Please upload sample-data.xls:'}

	return render(request, 'upload_forms/fr_upload.html', context)

def import_check(request):
	if request.method == "POST":
		form = UploadFileForm(request.POST,
							  request.FILES)
		if form.is_valid():
			request.FILES['file'].save_to_database(
				model=CheckProduct,
				mapdict=["sku_ns", "brand", "desc", "confirmed_placed"])
			messages.success(request, 'records successfully imported to database.')
			return HttpResponseRedirect('/main/')
		else:
			return HttpResponseBadRequest()
	else:
		form = UploadFileForm()
		context = {'form': form, 'title': 'Import excel data into database example', 'header': 'Please upload sample-data.xls:'}

	return render(request, 'upload_forms/fr_upload.html', context)

def export_studiocheck(request):

	records = Product.objects.all()
	warnrecords = []
	for record in records:
		if record.studio_out is None:
			continue
		elif all([record.confirmed_placed is None, record.studio_out <= (date - timedelta(days=14)), record.studio_out >= date - timedelta(days=160)]):
			if len(record.item_no) < 16:
				for item in records:
					if item.sku == record.item_no:
						warnrecords.append(item)
			else:
				warnrecords.append(record)

	query_sets = warnrecords
	column_names = ["ad_date", "vendor_number", "size_type",  "item_no", "item_ns", "mfg", "desc", "notes", "order_date", "received_dc", "received_buyer", "studio_out", "confirmed_placed", "received_other" ]
	return excel.make_response_from_query_sets(query_sets, column_names, 'xlsx')

def export_daily(request, atype):

	firstlinks = []
	vendrecords = []
	firstrecords = []

	if atype.startswith('0') or atype.startswith('1'):
		query_sets = Product.objects.filter(first_date__iexact=atype).exclude(Q(confirmed_placed__icontains='Jan') | Q(confirmed_placed__icontains='Feb') | Q(confirmed_placed__icontains='Mar') | Q(confirmed_placed__icontains='Apr') | Q(confirmed_placed__icontains='May') | Q(confirmed_placed__icontains='Jun') | Q(confirmed_placed__icontains='Jul') | Q(confirmed_placed__icontains='Aug') | Q(confirmed_placed__icontains='Sep') | Q(confirmed_placed__icontains='Oct') | Q(confirmed_placed__icontains='Nov') | Q(confirmed_placed__icontains='Dec')).exclude(received_buyer__isnull=False).exclude(received_dc__isnull=False).exclude(order_date__isnull=False).exclude(studio_out__isnull=False).exclude(studio_out__isnull=False).exclude(Q(dc_received_u__lt=200) & Q(curr_dc_oh_u__lt=200)).order_by('buyer')
		column_names = ["buyer", 'dc_received_u', 'curr_dc_oh_u', 'dc_curr_oo_u', "first_date", "size_type",  "item_ns", "vendor_number",  "item_no", "mfg", "desc", "color_desc", "have_image", "ad_date", "order_date"]
		return excel.make_response_from_query_sets(query_sets, column_names, 'xlsx')

	else:
		return HttpResponseRedirect('/first/')

def export_pickup(request, atype):

	firstlinks = []
	vendrecords = []
	firstrecords = []

	if atype.startswith('0') or atype.startswith('1'):
		# query_sets = AdProduct.objects.filter(Q(ad_date__iexact=atype) & Q(studio_out__iexact="None") & Q(received_buyer__iexact="None") & Q(received_dc__iexact="None") & Q(order_date__iexact="None")& (Q(confirmed_placed__isnull=True) | Q(confirmed_placed__iexact="None"))).exclude(Q(notes__icontains='DNS'))
		query_sets = AdProduct.objects.filter(ad_date__iexact=atype)
		# column_names = ["ad_date", "size_type", "item_ns", "vendor_number", "item_no",  "mfg", "desc", "notes"]
		column_names = ['curr_dc_oh_u', 'dc_curr_oo_u', "size_type",  "item_ns", "vendor_number",  "item_no", "mfg", "desc", "ad_date", "order_date"]

		return excel.make_response_from_query_sets(query_sets, column_names, 'xlsx')

	else:
		return HttpResponseRedirect('/adscreen/')

def export_required(request, atype):

	firstlinks = []
	vendrecords = []
	firstrecords = []

	if atype.startswith('0') or atype.startswith('1'):
		# query_sets = AdProduct.objects.filter(Q(ad_date__iexact=atype) & Q(studio_out__iexact="None") & Q(received_buyer__iexact="None") & Q(received_dc__iexact="None") & Q(order_date__iexact="None")& (Q(confirmed_placed__isnull=True) | Q(confirmed_placed__iexact="None"))).exclude(Q(notes__icontains='DNS'))
		query_sets = AdProduct.objects.filter(Q(ad_date__iexact=atype) & Q(confirmed_placed__isnull=True)).exclude(notes__icontains="DNS").exclude(received_buyer__isnull=False).exclude(received_dc__isnull=False).exclude(order_date__isnull=False).exclude(studio_out__isnull=False).exclude(curr_dc_oh_u__gte=1)
		column_names = ["ad_date", "vendor_number", "item_ns", "size_type", "item_no",  "mfg", "desc", "notes"]
		return excel.make_response_from_query_sets(query_sets, column_names, 'xlsx')

	else:
		return HttpResponseRedirect('/adscreen/')

def export_hotlist(request, atype):

	if atype:
		# query_sets = AdProduct.objects.filter(Q(ad_date__iexact=atype) & Q(studio_out__iexact="None") & Q(received_buyer__iexact="None") & Q(received_dc__iexact="None") & Q(order_date__iexact="None")& (Q(confirmed_placed__isnull=True) | Q(confirmed_placed__iexact="None"))).exclude(Q(notes__icontains='DNS'))
		query_sets = HotItem.objects.all().order_by('ad_date')
		column_names = ["create_date", "item_no", "ad_date", "item_name"]
		return excel.make_response_from_query_sets(query_sets, column_names, 'xlsx')
	else:
		return HttpResponseRedirect('/hotlist/')
   
def export_mods(request, atype):
# exports objects from replacedimage table by change date to an .xlsx file
	def formatter(input):
		# converts "010115" into "2015-01-15"
		inputstr = str(input)
		if len(inputstr) == 6:
			year = "20"+inputstr[-2:]
			month = inputstr[0:2]
			day = inputstr[2:4]
		try:
			return year + "-" + month + "-" + day
		except:
			return "none"

	target = formatter(atype)
	print(target)

	if target.startswith('2016'):
		query_sets = ReplacedImage.objects.filter(change_date__iexact=target)
		column_names = ["sku", 'sku_ns', 'item_no', 'item_ns', "old_filename",  "new_filename", "change_date"]
		return excel.make_response_from_query_sets(query_sets, column_names, 'xlsx')

	else:
		return HttpResponseRedirect('/b5/view_replaced')

def productsearch(request, searchstring):
# searchbox takes cvs and returns a page with a thumbnail and list of every product from input class on the export server    
	printrecords = PrintFile.objects.filter(assoc_sku__istartswith=searchstring)
	lgrecords = LargeWebfiles.objects.filter(long_sku__istartswith=searchstring)
	rgrecords = RegularWebfiles.objects.filter(long_sku__istartswith=searchstring)
	msrecords = MSWebfiles.objects.filter(long_sku__istartswith=searchstring)
	threcords = ThumbWebfiles.objects.filter(long_sku__istartswith=searchstring) 
	tmp = "class_search.html"
	variables = {'printrecords': printrecords, 'lgrecords': lgrecords, 'rgrecords': rgrecords, 'msrecords': msrecords, 'threcords': threcords }
	return render(request, tpl, variables)


def success_message(request):
	return redirect('success.html')


def replace_image(request):
# formset that adds records to replacedimage table
	def custom_field_callback(field):
		user = request.user
		profile = request.user.profile
		if field.name == 'processor':
			return field.formfield(initial=profile)
		else:
			return field.formfield()

	RIFormset = modelformset_factory(ReplacedImage, formfield_callback=custom_field_callback, fields='__all__', extra=5, form=RIForm)
	queryset = ReplacedImage.objects.none()
	formset = RIFormset(queryset=queryset)
	

	saved_records = 0
	if request.method == 'POST':
		formset = RIFormset(request.POST, request.FILES)

		if(formset.is_valid()):
			for form in formset:
				if form.has_changed():
					saved_records += 1
			formset.save()
			messages.success(request, '(' + str(saved_records) +') records successfully added to database.')
			return HttpResponseRedirect('/b5/replace_image/')
		else:
			form_errors = formset.errors
			context = {'formset': formset, 'form_errors': form_errors}
			return render(request, 'b5/replace_image_form.html', context)

	else:
		context = {'formset': formset}
		return render(request, 'b5/replace_image_form.html', context)


def view_replaced(request):
# searchbox returns all replacedimage objects or returns replacedimage objects by input date (eg."011515")
	def formatter(input):
		inputstr = str(input)
		if len(inputstr) == 6:
			year = "20"+inputstr[-2:]
			month = inputstr[0:2]
			day = inputstr[2:4]
		try:
			return year + "-" + month + "-" + day
		except:
			return "none"

	form = RIForm()
	imagelinks = []
	records = []
	monthlist = ['jan','feb','mar','apr','may','jun','jul','aug','sept','oct','nov','dec']

	show_results = True
	if 'query' in request.GET:
		show_results = True
		query = request.GET['query'].strip()

		if query == "all":
			form = PostsearchForm({'query' : query})
			context = { 'form': form, 'records': records }
			return render(request, 'b5/replace_template.html', context)

		elif query.lower() in monthlist:
			val = query.lower()
			month = monthlist.index(val) + 1
			if len(str(month)) == 1:
				month = "0" + str(month) 
			else:
				month = str(month)
			records = ReplacedImage.objects.filter(change_date__icontains="'-"+month+"-")
			form = PostsearchForm({'query' : query})
			context = { 'form': form, 'records': records }
			return render(request, 'b5/replace_template.html', context)

		elif query:
			form = PostsearchForm({'query' : query})
			target = formatter(query)
			records = ReplacedImage.objects.filter(change_date__iexact=target)
			context = { 'form': form, 'records': records }
			return render(request, 'b5/replace_template.html', context)


	else:
		form = PostsearchForm()
		tpl = 'b5/replace_template.html'
		variables = {'form':form }
		return render(request, tpl, variables)




def export_data(request, atype):
# exports objects to .xlsx worksheets based on supported query inputs
	monthlist = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'july', 'aug', 'sep', 'oct', 'nov', 'dec']
	alllinks = []
	worklinks = []
	firstlinks = []
	productlinks = []
	skulinks = []
	imagelinks = []
	buylinks = []
	checklinks = []
	adlinks = []
	dec_reportlinks = []

	vendrecords = []
	firstrecords = []


	if atype == "sheet":
		return excel.make_response_from_a_table(DexProduct, 'xlsx')

	elif atype in monthlist:
		target = str(monthlist.index(atype) + 1)
		if len(target) == 1:
			target = "0" + target
			print(target)
		query_sets = Product.objects.filter(ad_date__istartswith=target).order_by('ad_date')
		column_names = ["desc", 'vendor_number', 'ad_date', "confirmed_placed",  "item_no", "mfg",  "notes", "order_date", "received_dc", "studio_out", "received_buyer", "received_other", 'item_ns', "sku", "sku_ns", "first", "buyer",  "first_date",  "na"]
		return excel.make_response_from_query_sets(query_sets, column_names, 'xlsx')

	elif atype == "dec_report":
		vendrecords = Product.objects.filter(Q(first__exact='t') & Q(first_date__istartswith='12') & (Q(desc__icontains="NIKE") | Q(desc__icontains="UA") | Q(desc__icontains="UNDER"))).exclude(ad_date__isnull=False).exclude(confirmed_placed__isnull=True).exclude(confirmed_placed='').exclude(studio_out__isnull=False)
		firstrecords = Product.objects.filter(Q(ad_date__istartswith='12') & (Q(received_other__isnull=False) | Q(first__iexact='t')))
		result_list = list(chain(vendrecords, firstrecords))
		query_sets = result_list
		column_names = ['ad_date', 'vendor_number', "item_no", "mfg", "desc", "notes", "order_date", "received_dc", "received_137", "received_buyer", "studio_out", "confirmed_placed"]
		return excel.make_response_from_query_sets(query_sets, column_names, 'xlsx')

	elif atype == "jan_report":
		vendrecords = Product.objects.filter(Q(first__exact='t') & Q(first_date__istartswith='01') & (Q(desc__icontains="NIKE") | Q(desc__icontains="UA") | Q(desc__icontains="UNDER"))).exclude(ad_date__isnull=False).exclude(confirmed_placed__isnull=True).exclude(confirmed_placed='')
		firstrecords = Product.objects.filter(Q(ad_date__istartswith='01') & (Q(received_other__isnull=False) | Q(first__iexact='t'))).order_by('ad_date')
		result_list = list(chain(vendrecords, firstrecords))
		query_sets = result_list
		column_names = ['ad_date', 'vendor_number', "item_no", "mfg", "desc", "notes", "order_date", "received_dc", "received_137", "received_buyer", "studio_out", "confirmed_placed"]
		return excel.make_response_from_query_sets(query_sets, column_names, 'xlsx')

	elif atype == "feb_report":
		vendrecords = Product.objects.filter(Q(first__exact='t') & Q(first_date__istartswith='02') & (Q(desc__icontains="NIKE") | Q(desc__icontains="UA") | Q(desc__icontains="UNDER"))).exclude(ad_date__isnull=False).exclude(confirmed_placed__isnull=True).exclude(confirmed_placed='')
		firstrecords = Product.objects.filter(Q(ad_date__istartswith='02') & (Q(received_other__isnull=False) | Q(first__iexact='t'))).order_by('ad_date')
		result_list = list(chain(vendrecords, firstrecords))
		query_sets = result_list
		column_names = ['ad_date', 'vendor_number', "item_no", "mfg", "desc", "notes", "order_date", "received_dc", "received_137", "received_buyer", "studio_out", "confirmed_placed"]
		return excel.make_response_from_query_sets(query_sets, column_names, 'xlsx')

	elif atype == "febweb":
		query_sets = Product.objects.filter(Q(first_date__istartswith='02') & Q(confirmed_placed__isnull=True) & Q(studio_out__isnull=True)).exclude(Q(received_buyer__isnull=False) | Q(received_dc__isnull=False) | Q(first__exact='t'))
		column_names = ["desc", 'vendor_number', 'ad_date', "confirmed_placed",  "item_no", "mfg",  "notes", "order_date", "received_dc", "studio_out", "received_buyer", "received_other", 'item_ns', "sku", "sku_ns", "first", "buyer",  "first_date",  "na"]
		return excel.make_response_from_query_sets(query_sets, column_names, 'xlsx')

	elif atype == "janweb":
		query_sets = Product.objects.filter(Q(first_date__istartswith='01') & Q(confirmed_placed__isnull=True) & Q(studio_out__isnull=True)).exclude(Q(received_buyer__isnull=False) | Q(received_dc__isnull=False) | Q(first__exact='t'))
		column_names = ['ad_date', 'vendor_number', "item_no", "mfg", "desc", "notes", "order_date", "received_dc", "received_137", "received_buyer", "studio_out", "confirmed_placed"]
		return excel.make_response_from_query_sets(query_sets, column_names, 'xlsx')

	elif atype == "all":
		query_sets = Product.objects.all()
		column_names = ["desc", 'vendor_number', 'ad_date', "confirmed_placed",  "item_no", "mfg",  "notes", "order_date", "received_dc", "studio_out", "received_buyer", "received_other", 'item_ns', "sku", "sku_ns", "first", "buyer",  "first_date",  "na"]
		return excel.make_response_from_query_sets(query_sets, column_names, 'xlsx')

	elif atype == "unavailable":
		query_sets = Product.objects.filter(Q(na="t") | Q(na="true")) 
		column_names = ["desc", 'vendor_number', 'ad_date', "confirmed_placed",  "item_no", "mfg",  "notes", "order_date", "received_dc", "studio_out", "received_buyer", "received_other", 'item_ns', "sku", "sku_ns", "first", "buyer",  "first_date",  "na"]
		return excel.make_response_from_query_sets(query_sets, column_names, 'xlsx')

	elif atype == "fr":
		query_sets = FirstReceipt.objects.all()
		column_names = ["buyer", "dc_received_u", "curr_dc_oh_u", "dc_curr_oo_u", "date_received", "item_ns", "vendor_number", "item_no", "vendor_style", "description", "have_image", "ad_date", "order_date", "received_dc", "received_137", "from_file", "photo_dldate", "whowhen", "studio_out", "checked_out", "confirmed_placed", "studio_in", "merch_to_137"]
		return excel.make_response_from_query_sets(query_sets, column_names, 'xlsx')

	elif "daily-fr" in atype:
		query_sets = FirstReceipt.objects.filter(first_date__icontains=atype)
		column_names = ["buyer", "dc_received_u", "curr_dc_oh_u", "dc_curr_oo_u", "date_received", "item_ns", "short_sku", "item_no", "vendor_style", "description", "have_image", "ad_date", "order_date", "received_dc", "received_137", "from_file", "photo_dldate", "whowhen", "studio_out", "checked_out", "confirmed_placed", "studio_in", "merch_to_137"]
		return excel.make_response_from_query_sets(query_sets, column_names, 'xlsx')

	elif "vendor" in atype:
		query_sets = Product.objects.filter((Q(first_date__istartswith=cmonth) | Q(first_date__istartswith=cmonth) | (Q(ad_date__istartswith=cmonth) | Q(ad_date__istartswith=cmonth))) & Q(desc__icontains="NIKE")).exclude(na="t").exclude(confirmed_placed__isnull=False).exclude(studio_out__isnull=False).exclude(Q(received_buyer__isnull=False) | Q(received_dc__isnull=False))
		column_names = ["desc", 'vendor_number', 'ad_date', "confirmed_placed",  "item_no", "mfg",  "notes", "order_date", "received_dc", "studio_out", "received_buyer", "received_other", 'item_ns', "sku", "sku_ns", "first", "buyer",  "first_date",  "na"]
		return excel.make_response_from_query_sets(query_sets, column_names, 'xlsx')


	elif atype.startswith('0') or atype.startswith('1'):
		query_sets = Product.objects.filter(ad_date=atype)
		column_names = ["desc", 'vendor_number', 'ad_date', "confirmed_placed",  "item_no", "mfg",  "notes", "order_date", "received_dc", "studio_out", "received_buyer", "received_other", 'item_ns', "sku", "sku_ns", "first", "buyer",  "first_date",  "na"]
		return excel.make_response_from_query_sets(query_sets, column_names, 'xlsx')

	else:
		return HttpResponseRedirect('/main/')


def createEntry(request):
	if request.method == 'POST':
		form = ProductForm(request.POST, request.FILES)

		if form.is_valid():
			form.save()
			tpl = 'b5/Product_form.html'
			variables = RequestContext(request, {'form':form })
			return HttpResponseRedirect('/b5/create-entry/')
		
		else:
				form = ProductForm()
				tpl = 'b5/Product_form.html'
				variables = {'form':form }
				return render(request, tpl, variables)
	else:
		form = ProductForm()
		tpl = 'b5/Product_form.html'
		variables = {'form':form }
		return render(request, tpl, variables)


def updateEntry(request, id):
	if request.method == 'POST':
		form = ProductForm(request.POST, request.FILES, instance = Product.objects.get(pk = id))
		if request.POST.get('delete'):
			instance = Product.objects.get(pk=id)
			instance.delete()
			return HttpResponseRedirect('/main/')

		if form.is_valid():
			URL = request.POST.get('URL')
			form.save()
			return HttpResponseRedirect(URL)
		
	else:
		form = ProductForm(instance = Product.objects.get(pk = id))
		tpl = 'b5/postupdate_form.html'
		variables = {'form':form }
		return render(request, tpl, variables)



def multipost(request):
	ProductFormset = modelformset_factory(Product, exclude=('product_class',), extra=30, form=ProductForm)
	saved_records = 0
	if request.method == 'POST':
		formset = ProductFormset(request.POST, request.FILES)

		if(formset.is_valid()):
			for form in formset:
				if form.has_changed():
					saved_records += 1
			formset.save()
			messages.success(request, '(' + str(saved_records) +') records successfully added to database.')
			return HttpResponseRedirect('/b5/multipost/')

	else:
		queryset = Product.objects.none()
		formset = ProductFormset(queryset=queryset)
		variables = {'formset':formset }
		return render(request, "b5/multipost_form.html", variables)

		

def edit_multipost(request, ad_date):
	ProductFormset = modelformset_factory(Product, exclude=('product_class',), extra=30, form=ProductForm)
	saved_records = 0
	if request.method == 'POST':
		formset = ProductFormset(request.POST, request.FILES)

		if(formset.is_valid()):
			for form in formset:
				if form.has_changed():
					saved_records += 1
			formset.save()
			messages.success(request, '(' + str(saved_records) +') records successfully added to database.')
			return HttpResponseRedirect('/b5/multipost/')

	else:
		queryset = Product.objects.filter(ad_date=ad_date)
		formset = ProductFormset(queryset=queryset)
		variables = {'formset':formset }
		return render(request, "b5/multipost_form.html", variables)


def add_first(request):
	def custom_field_callback(field):
		if field.name == 'first_date':
			return field.formfield(required=True)
		if any([field.name == 'confirmed_placed', field.name == 'ad_date']):
			return field.formfield(initial='None')
		else:
			return field.formfield()

	FRFormset = modelformset_factory(Product, formfield_callback=custom_field_callback, exclude=('product_class',), extra=30, form=FRForm)
	queryset = Product.objects.none()
	formset = FRFormset(queryset=queryset)
	# variables = RequestContext(request, {'formset':formset })
	saved_records = 0
	if request.method == 'POST':
		formset = FRFormset(request.POST, request.FILES)

		if(formset.is_valid()):
			for form in formset:
				if form.has_changed():
					saved_records += 1
			formset.save()
			messages.success(request, '(' + str(saved_records) +') records successfully added to database.')
			return HttpResponseRedirect('/b5/add_first/')
		else:
			form_errors = formset.errors
			return render(request, 'b5/addfirst_form.html', {'formset': formset, 'form_errors': form_errors})

	else:
		return render(request, 'b5/addfirst_form.html', {'formset': formset})




def multipost_init(request, client, id):
	ProductFormset = modelformset_factory(Product, exclude=('product_class',), extra=6, formset=BaseProductFormSet)
	if request.method == 'POST':
		formset = ProductFormset(request.POST, request.FILES)
		
		if(formset.is_valid()):
			formset.save()
			return HttpResponseRedirect('/b5/multipost/')

	else:
		formset = ProductFormset(initial=[{'client': client, 'job_number': id,},{'client': client, 'job_number': id,},{'client': client, 'job_number': id,},{'client': client, 'job_number': id,},{'client': client, 'job_number': id,},{'client': client, 'job_number': id,}])
		clients = ClientList.objects.all()
		variables = {'clients' : clients, 'formset':formset }
		return render(request, "b5/multipost_form.html", variables)
 

class ProductUpdate(UpdateView):
	model = Product
	fields = '__all__'
	success_url = "../b5/newProduct/"


def export_dir(request):
	user = request.user
	path = '/Users/kevin/Desktop/archive/111715/large/'
	# path = '/Volumes/EXPORT/Color/large/'
	file_arr = os.listdir(path)
	records = {}

	for item in file_arr:
		date = os.stat(os.path.join(path + item)).st_mtime
		filetime = time.ctime(os.stat(os.path.join(path + item)).st_mtime)
		if item.startswith('.'):
			continue
		
		records[item] = filetime

	variables = { 'user': user, 'records': records } 
	return render(request, "b5/export.html", variables)


def edit_view(request):
  return HttpResponseRedirect("/admin/b5/product/")


def export_csv(qs, fields=None):
	
	qs = Product.objects.all()
	model = qs.model
	response = HttpResponse()
	response['Content-Disposition'] = 'attachment; filename=%s.csv' % slugify(model.__name__)
	writer = csv.writer(response)
	# Write headers to CSV file
	if fields:
		headers = fields
	else:
		headers = []
		for field in model._meta.fields:
			headers.append(field.name)
	writer.writerow(headers)
	# Write data to CSV file
	for obj in qs:
		row = []
		for field in headers:
			if field in headers:
				val = getattr(obj, field)
				if callable(val):
					val = val()
				row.append(val)
		writer.writerow(row)
	# Return CSV file to browser as download
	return response


	
@login_required(login_url='/') 
def postsearch(request):
		form = PostsearchForm() 


		alllinks = []
		worklinks = []
		firstlinks = []
		productlinks = []
		skulinks = []
		imagelinks = []
		monthlinks = []
		buylinks = []
		checklinks = []
		adlinks = []
		dec_reportlinks = []
		jan_reportlinks = []
		feb_reportlinks = []
		report_vendorlinks = []

		user = request.user
		records = []
		tmrecords = InboxEntry.objects.filter( Q(assigned_to=user.id) & Q(priority__iexact='urgent'))
		watchrecords = HotItem.objects.all()
		allwatchrecords = HotItem.objects.all()
		cgrecords = ColorGrid.objects.all()
		printrecords = []
		lgrecords = []
		rgrecords = []
		msrecords = []
		threcords = []
		monthlist = ['jan','feb','mar','apr','may','jun','jul','aug','sept','oct','nov','dec']
		hotlinks = []

		current_user = request.user
		id = current_user.id
		u = UserProfile.objects.get(pk=id)
		boxrecords = TaskBox.objects.filter(Q(subscribers__id__contains=id))
		boxes = boxrecords.count()
		boxdict = {}
		for x in range(1, boxes + 1):
			boxdict['box{0}s'.format(x)] = InboxEntry.objects.filter(box='{0}'.format(x))
		myrecords = InboxEntry.objects.filter(assigned_to=u).order_by('-priority', '-status')

		def returndata(data, default_dict, template):
			records = data
			record_dict = {'records': records}
			z = default_dict.copy()
			z.update(boxdict)
			z.update(record_dict)
			return template, z

		show_results = True
		if 'query' in request.GET:
				show_results = True

				query = request.GET['query'].strip()

				if query.lower() in monthlist:
					val = query.lower()
					month = monthlist.index(val) + 1
					if len(str(month)) == 1:
						month = "0" + str(month) 
					else:
						month = str(month)
					monthlinks = Product.objects.filter(ad_date__istartswith=month)
					a,b = returndata(monthlinks,{'user': user, 'myrecords': myrecords, 'boxrecords': boxrecords, 'watchrecords': watchrecords, 'allwatchrecords': allwatchrecords, 'cgrecords': cgrecords, 'tmrecords': tmrecords, 'form': form}, 'list_template.html')
					return render(request, b, a)

				elif query:
						form = PostsearchForm({'query' : query})
						productlinks = Product.objects.filter( Q(item_no__iexact=query) | Q(item_ns__iexact=query) )
						skulinks = Product.objects.filter( Q(sku__iexact=query) | Q(sku_ns__iexact=query) )
						imagelinks = ThumbWebfiles.objects.filter(sku__iexact=query)
						buylinks = Product.objects.filter(buyer__icontains=query)
						adlinks = Product.objects.filter(ad_date__iexact=query)
						if " " in query:
							functional_sku = query[0:15]
						else:
							functional_sku = query[0:4] + " " + query[4:9] + " " + query[9:13]



		if    len(checklinks) >= 1: 
				records = checklinks
				printrecords = PrintFile.objects.all()
				context = {'user': user, 'records': records, 'cgrecords': cgrecords, 'printrecords': printrecords, 'form': form}
				return render(request, 'check_template.html', context)


		elif    len(adlinks) >= 1:
				records = adlinks
				default_dict = {'user': user, 'myrecords': myrecords, 'boxrecords': boxrecords, 'records': records, 'watchrecords': watchrecords, 'allwatchrecords': allwatchrecords, 'cgrecords': cgrecords, 'tmrecords': tmrecords, 'form': form}
				z = default_dict.copy()
				z.update(boxdict)
				return render(request, 'list_template.html', z)


		elif    len(buylinks) >= 1:
				records = Product.objects.filter(buyer__icontains=query)
				default_dict = {'user': user, 'myrecords': myrecords, 'boxrecords': boxrecords, 'records': records, 'watchrecords': watchrecords, 'allwatchrecords': allwatchrecords, 'cgrecords': cgrecords, 'tmrecords': tmrecords, 'form': form}
				z = default_dict.copy()
				z.update(boxdict)
				return render(request, 'list_template.html', z)


		elif    len(dec_reportlinks) >= 1:
				records = Product.objects.filter(Q(first__exact='t') & Q(first_date__istartswith='12') & (Q(desc__icontains="NIKE") | Q(desc__icontains="UA") | Q(desc__icontains="UNDER"))).exclude(ad_date__isnull=False).exclude(confirmed_placed__isnull=True).exclude(confirmed_placed='').exclude(studio_out__isnull=False)
				workrecords = Product.objects.filter(Q(ad_date__istartswith='12') & (Q(received_other__isnull=False) | Q(first__iexact='t')))
				default_dict = {'user': user, 'myrecords': myrecords, 'boxrecords': boxrecords, 'records': records, 'watchrecords': watchrecords, 'allwatchrecords': allwatchrecords, 'workrecords': workrecords, 'cgrecords': cgrecords, 'tmrecords': tmrecords, 'form': form}
				z = default_dict.copy()
				z.update(boxdict)
				return render(request, 'list_template.html', z)

		elif    len(jan_reportlinks) >= 1:
				records = Product.objects.filter(Q(first__exact='t') & Q(first_date__istartswith='01') & (Q(desc__icontains="NIKE") | Q(desc__icontains="UA") | Q(desc__icontains="UNDER"))).exclude(ad_date__isnull=False).exclude(confirmed_placed__isnull=True).exclude(confirmed_placed='').exclude(studio_out__isnull=False)
				workrecords = Product.objects.filter(Q(ad_date__istartswith='01') & (Q(received_other__isnull=False) | Q(first__iexact='t')))
				default_dict = {'user': user, 'myrecords': myrecords, 'boxrecords': boxrecords, 'records': records, 'watchrecords': watchrecords, 'allwatchrecords': allwatchrecords, 'workrecords': workrecords, 'cgrecords': cgrecords, 'tmrecords': tmrecords, 'form': form}
				z = default_dict.copy()
				z.update(boxdict)
				return render(request, 'list_template.html', z)

		elif    len(feb_reportlinks) >= 1:
				records = Product.objects.filter(Q(first__exact='t') & Q(first_date__istartswith='02') & (Q(desc__icontains="NIKE") | Q(desc__icontains="UA") | Q(desc__icontains="UNDER"))).exclude(ad_date__isnull=False).exclude(confirmed_placed__isnull=True).exclude(confirmed_placed='').exclude(studio_out__isnull=False)
				workrecords = Product.objects.filter(Q(ad_date__istartswith='02') & (Q(received_other__isnull=False) | Q(first__iexact='t')))
				default_dict = {'user': user, 'myrecords': myrecords, 'boxrecords': boxrecords, 'records': records, 'watchrecords': watchrecords, 'allwatchrecords': allwatchrecords, 'workrecords': workrecords, 'cgrecords': cgrecords, 'tmrecords': tmrecords, 'form': form}
				z = default_dict.copy()
				z.update(boxdict)
				return render(request, 'list_template.html', z)

		
		elif    len(skulinks) >= 1:
				records = Product.objects.filter( Q(sku__iexact=query) | Q(sku_ns__iexact=query))
				printrecords = PrintFile.objects.filter(assoc_sku__iexact=functional_sku)
				lgrecords = LargeWebfiles.objects.filter( Q(long_sku__iexact=query) | Q(item_ns__iexact=query)  | Q(sku__iexact=query) | Q(sku_ns__iexact=query))
				rgrecords = RegularWebfiles.objects.filter( Q(long_sku__iexact=query) | Q(item_ns__iexact=query) | Q(sku__iexact=query) | Q(sku_ns__iexact=query))
				msrecords = MSWebfiles.objects.filter( Q(long_sku__iexact=query) | Q(item_ns__iexact=query) | Q(sku__iexact=query) | Q(sku_ns__iexact=query))
				threcords = ThumbWebfiles.objects.filter( Q(long_sku__iexact=query) | Q(item_ns__iexact=query) | Q(sku__iexact=query) | Q(sku_ns__iexact=query))
				display_dict = {'user': user, 'records': records, 'lgrecords': lgrecords, 'rgrecords': rgrecords, 'msrecords': msrecords, 'threcords': threcords, 'printrecords': printrecords, 'form': form}
				return render(request, 'display_template.html', display_dict)


		elif    len(productlinks) >= 1:
				records = Product.objects.filter( Q(item_no__iexact=query) | Q(item_ns__iexact=query))
				printrecords = PrintFile.objects.filter(assoc_sku__iexact=functional_sku)
				lgrecords = LargeWebfiles.objects.filter( Q(long_sku__iexact=query) | Q(item_ns__iexact=query)  | Q(sku__icontains=query))
				rgrecords = RegularWebfiles.objects.filter( Q(long_sku__iexact=query) | Q(item_ns__iexact=query) | Q(sku__iexact=query))
				msrecords = MSWebfiles.objects.filter( Q(long_sku__iexact=query) | Q(item_ns__iexact=query) | Q(sku__iexact=query))
				threcords = ThumbWebfiles.objects.filter( Q(long_sku__iexact=query) | Q(item_ns__iexact=query) | Q(sku__icontains=query) )
				display_dict = {'user': user, 'records': records, 'lgrecords': lgrecords, 'rgrecords': rgrecords, 'msrecords': msrecords, 'threcords': threcords, 'printrecords': printrecords, 'form': form}
				return render(request, 'display_template.html', display_dict)


		elif    len(imagelinks) >= 1:
				records = Product.objects.filter( Q(item_no__iexact=query) | Q(item_ns__iexact=query) | Q(sku__iexact=query) )
				printrecords = PrintFile.objects.filter(assoc_sku__iexact=query)
				lgrecords = LargeWebfiles.objects.filter( Q(long_sku__iexact=query) | Q(item_ns__iexact=query)  | Q(sku__iexact=query))
				rgrecords = RegularWebfiles.objects.filter( Q(long_sku__iexact=query) | Q(item_ns__iexact=query) | Q(sku__iexact=query))
				msrecords = MSWebfiles.objects.filter( Q(long_sku__iexact=query) | Q(item_ns__iexact=query) | Q(sku__iexact=query))
				threcords = ThumbWebfiles.objects.filter( Q(long_sku__iexact=query) | Q(item_ns__iexact=query) | Q(sku__iexact=query) )
				display_dict = {'user': user, 'records': records, 'lgrecords': lgrecords, 'rgrecords': rgrecords, 'msrecords': msrecords, 'threcords': threcords, 'printrecords': printrecords, 'form': form}
				return render(request, 'display_template.html', display_dict)


		else:
				default_dict = {'user': user, 'myrecords': myrecords, 'boxrecords': boxrecords, 'form': form, 'allwatchrecords': allwatchrecords, 'watchrecords': watchrecords, 'tmrecords': tmrecords, 'show_results': show_results} 
				z = default_dict.copy()
				z.update(boxdict)
				return render(request, 'list_template.html', z)

@login_required(login_url='/') 
def frsearch(request):
		form = PostsearchForm() 

		firstlinks = []
		alllinks = []
		productlinks = []
		skulinks = []
		imagelinks = []
		
		user = request.user
		records = []
		tmrecords = []
		watchrecords = []
		cgrecords = []
		printrecords = []
		lgrecords = []
		rgrecords = []
		msrecords = []
		threcords = []

		show_results = True
		if 'query' in request.GET:
				show_results = True
				query = request.GET['query'].strip()
				if query == "all":
					alllinks = Product.objects.filter(Q(first_date__isnull=False) | Q(first_date__exact=""))
				elif query:
						form = PostsearchForm({'query' : query})
						firstlinks = Product.objects.filter(first_date__iexact=query)

		if    len(alllinks) >= 1:
				records = Product.objects.filter(Q(first_date__isnull=False) | Q(first_date__exact="")).order_by('item_no')
				tmrecords = InboxEntry.objects.filter( Q(assigned_to=user.id) & Q(priority__iexact='urgent') )
				watchrecords = HotItem.objects.all()
				cgrecords = ColorGrid.objects.all()
				default_dict = {'user': user, 'records': records, 'watchrecords': watchrecords, 'cgrecords': cgrecords, 'tmrecords': tmrecords, 'form': form}
				return render(request, 'fr_list_template.html', default_dict)

		elif    len(firstlinks) >= 1:
				records = Product.objects.filter(first_date__iexact=query).order_by('item_no')
				watchrecords = HotItem.objects.all()
				tmrecords = InboxEntry.objects.filter( Q(assigned_to=user.id) & Q(priority__iexact='urgent') )
				cgrecords = ColorGrid.objects.all()
				default_dict = {'user': user, 'records': records, 'watchrecords': watchrecords, 'cgrecords': cgrecords, 'tmrecords': tmrecords, 'form': form}
				return render(request, 'fr_list_template.html', default_dict)

		else:
				
				# profile = UserProfile(user=user)
				watchrecords = HotItem.objects.all()
				tmrecords = InboxEntry.objects.filter( Q(assigned_to=user.id) & Q(priority__iexact='urgent') )
				tpl = "fr_list_template.html"
				variables = {'user': user, 'form': form, 'watchrecords': watchrecords, 'tmrecords': tmrecords,
				'show_results': show_results} 
				return render(request, tpl, variables)

@login_required(login_url='/') 
def adscreen(request):
		form = PostsearchForm() 

		alllinks = []
		worklinks = []
		firstlinks = []
		productlinks = []
		skulinks = []
		imagelinks = []
		monthlinks = []
		buylinks = []
		checklinks = []
		adlinks = []
		dec_reportlinks = []
		jan_reportlinks = []
		feb_reportlinks = []
		report_vendorlinks = []
		noimagelinks = []
		availablelinks = []
		unavailablelinks = []
		orderlinks = []

		user = request.user
		records = []
		adrecords = AdProduct.objects.all().values('ad_date','version').distinct().order_by('ad_date').reverse()
		tmrecords = InboxEntry.objects.filter( Q(assigned_to=user.id) & Q(priority__iexact='urgent'))
		watchrecords = HotItem.objects.all()
		allwatchrecords = HotItem.objects.all()
		cgrecords = ColorGrid.objects.all()
		printrecords = []
		lgrecords = []
		rgrecords = []
		msrecords = []
		threcords = []
		monthlist = ['jan','feb','mar','apr','may','jun','jul','aug','sept','oct','nov','dec']
		hotlinks = []

		current_user = request.user
		id = current_user.id
		u = UserProfile.objects.get(pk=id)
		boxrecords = TaskBox.objects.filter(Q(subscribers__id__contains=id))
		boxes = boxrecords.count()
		boxdict = {}
		for x in range(1, boxes + 1):
			boxdict['box{0}s'.format(x)] = InboxEntry.objects.filter(box='{0}'.format(x))
		myrecords = InboxEntry.objects.filter(assigned_to=u).order_by('-priority', '-status')

		def returndata(data, default_dict, template):
			records = data
			record_dict = {'records': records}
			z = default_dict.copy()
			z.update(boxdict)
			z.update(record_dict)
			return template, z

		adlinks = []

		show_results = True
		if 'query' in request.GET:
				show_results = True
				query = request.GET['query'].strip()
				if query == "all":
					alllinks = Product.objects.filter(Q(first_date__isnull=False) | Q(first_date__exact=""))
				elif query in monthlist:
					target = monthlist.index(query) + 1
					if target < 10:
						target = "0" + str(target)
					target = str(target)
					form = PostsearchForm({'query' : query})
					adlinks = AdProduct.objects.filter(ad_date__istartswith=target)
				elif query.endswith('-t'):
					query = query.split(' ')
					query = query[0]
					form = PostsearchForm({'query' : query})
					noimagelinks = AdProduct.objects.filter(ad_date__istartswith=query).exclude(confirmed_placed__isnull=False)
				elif query.endswith('-a'):
					query = query.split(' ')
					query = query[0]
					form = PostsearchForm({'query' : query})
					availablelinks = AdProduct.objects.filter(ad_date__istartswith=query).exclude(confirmed_placed__isnull=False).exclude(Q(curr_dc_oh_u=0) & Q(merch_to_137=0)).exclude(na__iexact="t")
				elif query.endswith('-u'):
					query = query.split(' ')
					query = query[0]
					form = PostsearchForm({'query' : query})
					unavailablelinks = AdProduct.objects.filter(ad_date__istartswith=query).exclude(confirmed_placed__isnull=False).exclude(Q(curr_dc_oh_u__gt=0) & Q(merch_to_137__gt=0))
				elif query.endswith('-o'):
					query = query.split(' ')
					query = query[0]
					form = PostsearchForm({'query' : query})
					orderlinks = AdProduct.objects.filter(ad_date__istartswith=query).exclude(confirmed_placed__isnull=False).exclude(Q(dc_curr_oo_u=0))
				elif query:
					form = PostsearchForm({'query' : query})
					adlinks = AdProduct.objects.filter(ad_date__iexact=query)
					productlinks = AdProduct.objects.filter( Q(item_no__iexact=query) | Q(item_ns__iexact=query) )
					skulinks = AdProduct.objects.filter( Q(sku__iexact=query) | Q(sku_ns__iexact=query) )

		if    len(noimagelinks) >= 1:
				records = noimagelinks
				default_dict = {'user': user, 'myrecords': myrecords, 'adrecords': adrecords, 'boxrecords': boxrecords, 'records': records, 'watchrecords': watchrecords, 'allwatchrecords': allwatchrecords, 'cgrecords': cgrecords, 'tmrecords': tmrecords, 'form': form}
				z = default_dict.copy()
				z.update(boxdict)
				return render(request, 'adscreen_list_template.html', z)  

		elif    len(availablelinks) >= 1:
				records = availablelinks
				default_dict = {'user': user, 'myrecords': myrecords, 'adrecords': adrecords, 'boxrecords': boxrecords, 'records': records, 'watchrecords': watchrecords, 'allwatchrecords': allwatchrecords, 'cgrecords': cgrecords, 'tmrecords': tmrecords, 'form': form}
				z = default_dict.copy()
				z.update(boxdict)
				return render(request, 'adscreen_list_template.html', z)

		elif    len(unavailablelinks) >= 1:
				records = unavailablelinks
				default_dict = {'user': user, 'myrecords': myrecords, 'adrecords': adrecords, 'boxrecords': boxrecords, 'records': records, 'watchrecords': watchrecords, 'allwatchrecords': allwatchrecords, 'cgrecords': cgrecords, 'tmrecords': tmrecords, 'form': form}
				z = default_dict.copy()
				z.update(boxdict)
				return render(request, 'adscreen_list_template.html', z)

		elif    len(orderlinks) >= 1:
				records = orderlinks
				default_dict = {'user': user, 'myrecords': myrecords, 'adrecords': adrecords, 'boxrecords': boxrecords, 'records': records, 'watchrecords': watchrecords, 'allwatchrecords': allwatchrecords, 'cgrecords': cgrecords, 'tmrecords': tmrecords, 'form': form}
				z = default_dict.copy()
				z.update(boxdict)
				return render(request, 'adscreen_list_template.html', z)

		elif    len(adlinks) >= 1:
				records = adlinks
				default_dict = {'user': user, 'myrecords': myrecords, 'adrecords': adrecords, 'boxrecords': boxrecords, 'records': records, 'watchrecords': watchrecords, 'allwatchrecords': allwatchrecords, 'cgrecords': cgrecords, 'tmrecords': tmrecords, 'form': form}
				z = default_dict.copy()
				z.update(boxdict)
				return render(request, 'adscreen_list_template.html', z)

		elif    len(skulinks) >= 1:
				records = skulinks
				default_dict = {'user': user, 'myrecords': myrecords, 'adrecords': adrecords, 'boxrecords': boxrecords, 'records': records, 'watchrecords': watchrecords, 'allwatchrecords': allwatchrecords, 'cgrecords': cgrecords, 'tmrecords': tmrecords, 'form': form}
				z = default_dict.copy()
				z.update(boxdict)
				return render(request, 'adscreen_list_template.html', z)

		elif    len(productlinks) >= 1:
				records = productlinks
				default_dict = {'user': user, 'myrecords': myrecords, 'adrecords': adrecords, 'boxrecords': boxrecords, 'records': records, 'watchrecords': watchrecords, 'allwatchrecords': allwatchrecords, 'cgrecords': cgrecords, 'tmrecords': tmrecords, 'form': form}
				z = default_dict.copy()
				z.update(boxdict)
				return render(request, 'adscreen_list_template.html', z)

		else:
				
				# profile = UserProfile(user=user)
				watchrecords = HotItem.objects.all()
				tmrecords = InboxEntry.objects.filter( Q(assigned_to=user.id) & Q(priority__iexact='urgent') )
				tpl = "adscreen_list_template.html"
				variables = {'user': user, 'form': form, 'adrecords': adrecords, 'watchrecords': watchrecords, 'tmrecords': tmrecords,
				'show_results': show_results}
				return render(request, tpl, variables)



@login_required(login_url='/') 
def pickupsearch(request):
	pass
		# form = PostsearchForm() 


		# alllinks = []
		# productlinks = []
		# decweblinks = []
		# janweblinks = []
		# febweblinks = []
		# marweblinks = []
		# declinks = []
		# janlinks = []
		# feblinks = []
		# marlinks = []
		# adlinks = []

		# user = request.user
		# records = []
		# tmrecords = []
		# watchrecords = []
		# cgrecords = []


		# show_results = True
		# if 'query' in request.GET:
		# 		show_results = True
		# 		query = request.GET['query'].strip()
		# 		if query == "all":
		# 			alllinks = Product.objects.filter(Q(confirmed_placed__isnull=True) & Q(studio_out__isnull=True)).exclude(Q(received_buyer__isnull=False) | Q(received_dc__isnull=False) | Q(first__exact='f'))

		# 		elif query == "jan":
		# 			janlinks = Product.objects.filter(Q(ad_date__istartswith='02') & Q(received_dc__isnull=False))

		# 		elif query == "feb":
		# 			feblinks = Product.objects.filter(Q(ad_date__istartswith='02') & Q(confirmed_placed__isnull=True) & Q(studio_out__isnull=True)).exclude(Q(received_buyer__isnull=False) | Q(received_dc__isnull=False) | Q(first__exact='f'))

		# 		elif query == "marweb":
		# 			marweblinks = Product.objects.filter(Q(first_date__istartswith='03') & Q(confirmed_placed__isnull=True) & Q(studio_out__isnull=True)).exclude(Q(received_buyer__isnull=False) | Q(received_dc__isnull=False) | Q(first__exact='t'))

		# 		elif query == "decweb":
		# 			decweblinks = Product.objects.filter(Q(first_date__istartswith='03') & Q(confirmed_placed__isnull=True) & Q(studio_out__isnull=True)).exclude(Q(received_buyer__isnull=False) | Q(received_dc__isnull=False) | Q(first__exact='t'))

		# 		elif query == "janweb":
		# 			# janweblinks = Product.objects.filter(Q(first_date__istartswith='1') & Q(confirmed_placed__isnull=True) & Q(studio_out__isnull=True)).exclude(Q(received_buyer__isnull=False) | Q(received_dc__isnull=False) | Q(first__exact='t'))
		# 			janweblinks = Product.objects.filter(Q(first_date__istartswith='01') & Q(order_date__isnull=False))

		# 		elif query == "febweb":
		# 			febweblinks = Product.objects.filter(Q(first_date__istartswith='02') & Q(confirmed_placed__isnull=True) & Q(studio_out__isnull=True)).exclude(Q(received_buyer__isnull=False) | Q(received_dc__isnull=False) | Q(first__exact='t'))


		# 		elif query:
		# 				form = PostsearchForm({'query' : query})
		# 				adlinks = Product.objects.filter(Q(ad_date__iexact=query) & Q(confirmed_placed__isnull=True) & Q(studio_out__isnull=True)).exclude(Q(received_buyer__isnull=False) | Q(received_dc__isnull=False) | Q(first__exact='t'))



		# if    len(janlinks) >= 1:
		# 		records = Product.objects.filter(Q(ad_date__istartswith='02') & Q(received_dc__isnull=False))
		# 		tmrecords = InboxEntry.objects.filter( Q(assigned_to=user.id) & Q(priority__iexact='urgent'))
		# 		watchrecords = HotItem.objects.all()
		# 		cgrecords = ColorGrid.objects.all()
		# 		default_dict = {'user': user, 'records': records, 'watchrecords': watchrecords, 'cgrecords': cgrecords, 'tmrecords': tmrecords, 'form': form}
		# 		return render(request, 'pickup_template.html', default_dict)

		# elif    len(feblinks) >= 1:
		# 		records = Product.objects.filter(Q(ad_date__istartswith='02') & Q(confirmed_placed__isnull=True) & Q(studio_out__isnull=True)).exclude(Q(received_buyer__isnull=False) | Q(received_dc__isnull=False) | Q(first__exact='f'))
		# 		tmrecords = InboxEntry.objects.filter( Q(assigned_to=user.id) & Q(priority__iexact='urgent') )
		# 		watchrecords = HotItem.objects.all()
		# 		cgrecords = ColorGrid.objects.all()
		# 		default_dict = {'user': user, 'records': records, 'watchrecords': watchrecords, 'cgrecords': cgrecords, 'tmrecords': tmrecords, 'form': form}
		# 		return render(request, 'pickup_template.html', default_dict)

		# elif    len(decweblinks) >= 1:
		# 		records = Product.objects.filter(Q(first_date__istartswith='12') & Q(confirmed_placed__isnull=True) & Q(studio_out__isnull=True)).exclude(Q(received_buyer__isnull=False) | Q(received_dc__isnull=False) | Q(first__exact='t'))
		# 		tmrecords = InboxEntry.objects.filter( Q(assigned_to=user.id) & Q(priority__iexact='urgent') )
		# 		watchrecords = HotItem.objects.all()
		# 		cgrecords = ColorGrid.objects.all()
		# 		default_dict = {'user': user, 'records': records, 'watchrecords': watchrecords, 'cgrecords': cgrecords, 'tmrecords': tmrecords, 'form': form}
		# 		return render(request, 'pickup_template.html', default_dict)

		# elif    len(marweblinks) >= 1:
		# 		records = Product.objects.filter(Q(first_date__istartswith='03') & Q(confirmed_placed__isnull=True) & Q(studio_out__isnull=True)).exclude(Q(received_buyer__isnull=False) | Q(received_dc__isnull=False) | Q(first__exact='t'))
		# 		tmrecords = InboxEntry.objects.filter( Q(assigned_to=user.id) & Q(priority__iexact='urgent') )
		# 		watchrecords = HotItem.objects.all()
		# 		cgrecords = ColorGrid.objects.all()
		# 		default_dict = {'user': user, 'records': records, 'watchrecords': watchrecords, 'cgrecords': cgrecords, 'tmrecords': tmrecords, 'form': form}
		# 		return render(request, 'pickup_template.html', default_dict)

		# elif    len(janweblinks) >= 1:
		# 		# records = Product.objects.filter(Q(first_date__istartswith='01') & Q(order_date__isnull=False) & (Q(received_dc__isnull=False) | Q(received_buyer__isnull=False))).exclude(Q(studio_out__isnull=False) | Q(confirmed_placed__isnull=False))
		# 		records = Product.objects.filter(Q(first_date__istartswith='01') & Q(order_date__isnull=False)).exclude(Q(studio_out__isnull=False) | Q(confirmed_placed__isnull=False))
		# 		tmrecords = InboxEntry.objects.filter( Q(assigned_to=user.id) & Q(priority__iexact='urgent') )
		# 		watchrecords = HotItem.objects.all()
		# 		cgrecords = ColorGrid.objects.all()
		# 		default_dict = {'user': user, 'records': records, 'watchrecords': watchrecords, 'cgrecords': cgrecords, 'tmrecords': tmrecords, 'form': form}
		# 		return render(request, 'pickup_template.html', default_dict)

		# elif    len(febweblinks) >= 1:
		# 		records = Product.objects.filter(Q(first_date__istartswith='02') & Q(confirmed_placed__isnull=True) & Q(studio_out__isnull=True)).exclude(Q(received_buyer__isnull=False) | Q(received_dc__isnull=False) | Q(first__exact='t'))
		# 		tmrecords = InboxEntry.objects.filter( Q(assigned_to=user.id) & Q(priority__iexact='urgent') )
		# 		watchrecords = HotItem.objects.all()
		# 		cgrecords = ColorGrid.objects.all()
		# 		default_dict = {'user': user, 'records': records, 'watchrecords': watchrecords, 'cgrecords': cgrecords, 'tmrecords': tmrecords, 'form': form}
		# 		return render(request, 'pickup_template.html', default_dict)

		# else:
				
		# 		# profile = UserProfile(user=user)
		# 		watchrecords = HotItem.objects.all()
		# 		tmrecords = InboxEntry.objects.filter( Q(assigned_to=user.id) & Q(priority__iexact='urgent') )
		# 		tpl = "pickup_template.html"
		# 		variables = {'user': user, 'form': form, 'watchrecords': watchrecords, 'tmrecords': tmrecords,
		# 		'show_results': show_results} 
		# 		return render(request, tpl, variables)


@login_required(login_url='/') 
def rumbasearch(request):
		form = PostsearchForm() 


		alllinks = []
		productlinks = []
		skulinks = []
	
		user = request.user
		records = []
		tmrecords = InboxEntry.objects.filter( Q(assigned_to=user.id) & Q(priority__iexact='urgent'))
		watchrecords = HotItem.objects.all()
		allwatchrecords = HotItem.objects.all()
		cgrecords = ColorGrid.objects.all()
		printrecords = []        
		hotlinks = []

		current_user = request.user
		id = current_user.id
		u = UserProfile.objects.get(pk=id)
		boxrecords = TaskBox.objects.filter(Q(subscribers__id__contains=id))
		boxes = boxrecords.count()
		boxdict = {}
		for x in range(1, boxes + 1):
			boxdict['box{0}s'.format(x)] = InboxEntry.objects.filter(box='{0}'.format(x))
		myrecords = InboxEntry.objects.filter(assigned_to=u).order_by('-priority', '-status')

		def returndata(data, default_dict, template):
			records = data
			record_dict = {'records': records}
			z = default_dict.copy()
			z.update(boxdict)
			z.update(record_dict)
			return template, z

		show_results = True
		if 'query' in request.GET:
				show_results = True

				query = request.GET['query'].strip()
				if len(query) > 19:
					temp_productlinks = []
					temp_skulinks = []
					span = 3
					words = str(query).split(" ")
					query = [" ".join(words[i:i+span]) for i in range(0, len(words), span)]
					# query = query[2]
					nquery = request.POST.getlist('query')
					form = PostsearchForm({'query' : query})
					for product in nquery:
						productlinks = RumbaProduct.objects.filter( Q(item_no__iexact=query) | Q(item_ns__iexact=query) )
						skulinks = RumbaProduct.objects.filter( Q(sku__iexact=product) | Q(sku_ns__iexact=product) )


					# if " " in query:
					#     functional_sku = query[0:15]
					# else:
					#     functional_sku = query[0:4] + " " + query[4:9] + " " + query[9:13]

				elif all([len(query) == 10, " " not in query]) or all([len(query) == 7, " " not in query]):
						form = PostsearchForm({'query' : query})
						productlinks = RumbaProduct.objects.filter(vendor_number__icontains=query)



				else:
						form = PostsearchForm({'query' : query})
						productlinks = RumbaProduct.objects.filter( Q(item_no__iexact=query) | Q(item_ns__iexact=query) )
						skulinks = RumbaProduct.objects.filter( Q(sku__iexact=query) | Q(sku_ns__iexact=query) )
						
						if " " in query:
							functional_sku = query[0:15]
						else:
							functional_sku = query[0:4] + " " + query[4:9] + " " + query[9:13]



		if    len(productlinks) >= 1:
				records = productlinks
				display_dict = {'user': user, 'records': records, 'form': form}
				return render(request, 'rumba_template.html', display_dict)


		elif    len(skulinks) >= 1:
				records = skulinks
				display_dict = {'user': user, 'records': records, 'form': form}
				return render(request, 'rumba_template.html', display_dict)


		else:
				default_dict = {'user': user, 'myrecords': myrecords, 'boxrecords': boxrecords, 'form': form, 'watchrecords': watchrecords, 'allwatchrecords': allwatchrecords, 'tmrecords': tmrecords, 'show_results': show_results} 
				z = default_dict.copy()
				z.update(boxdict)
				return render(request, 'rumba_template.html', z)


def inventory_819(request):
    form = PostsearchForm
    current_user = request.user
    id = current_user.id
    u = UserProfile.objects.get(pk=id)
    boxrecords = TaskBox.objects.filter(Q(subscribers__id__contains=id))
    boxes = boxrecords.count()
    boxdict = {}
    for x in range(1, boxes + 1):
        boxdict['box{0}s'.format(x)] = InboxEntry.objects.filter(box='{0}'.format(x))
    myrecords = InboxEntry.objects.filter(assigned_to=u).order_by('-priority', '-status')

    def returndata(data, default_dict, template):
        records = data
        record_dict = {'records': records}
        z = default_dict.copy()
        z.update(boxdict)
        z.update(record_dict)
        return template, z

    sourcelinks = []
    targetlinks = []
    productlinks = []
    worksheetrecords = InventoryProduct.objects.all().values('source').distinct()

    show_results = True
    if 'query' in request.GET:
            show_results = True
            query = request.GET['query'].strip()
            if len(query) == 7:
                form = PostsearchForm({'query' : query})
                targetlinks = InventoryProduct.objects.filter(sku__iexact=query)

            elif query:
                form = PostsearchForm({'query' : query})
                productlinks = InventoryProduct.objects.filter(item_no__iexact=query)
                sourcelinks = InventoryProduct.objects.filter(source__iexact=query)

    if    len(targetlinks) >= 1:
            records = targetlinks
            default_dict = {'current_user': current_user, 'boxrecords': boxrecords, 'records': records, 'form': form, 'worksheetrecords': worksheetrecords}
            z = default_dict.copy()
            z.update(boxdict)
            return render(request, '819_template.html', z)  

    elif    len(productlinks) >= 1:
            records = productlinks
            default_dict = {'current_user': current_user, 'boxrecords': boxrecords, 'records': records, 'form': form, 'worksheetrecords': worksheetrecords}
            z = default_dict.copy()
            z.update(boxdict)
            return render(request, '819_template.html', z)

    elif    len(sourcelinks) >= 1:
            records = sourcelinks
            default_dict = {'current_user': current_user, 'boxrecords': boxrecords, 'records': records, 'form': form, 'worksheetrecords': worksheetrecords}
            z = default_dict.copy()
            z.update(boxdict)
            return render(request, '819_template.html', z)

    else:
            tpl = "819_template.html"
            variables = {'current_user': current_user, 'form': form, 'worksheetrecords': worksheetrecords,
            'show_results': show_results}
            return render(request, tpl, variables)


def export_xlsx(modeladmin, request, queryset):
	import openpyxl
	from openpyxl.cell import get_column_letter
	response = HttpResponse(mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = 'attachment; filename=mymodel.xlsx'
	wb = openpyxl.Workbook()
	ws = wb.get_active_sheet()
	ws.title = "LatestTracking"

	row_num = 0

	columns = [
		(u"Ad Date", 70),
		(u"Item Number", 70),
		(u"Manufacturer", 70),
		(u"Description", 160),
		(u"Vendor Number", 70),
		(u"Order Date", 70),
		(u"Received DC", 70),
		(u"Received Buyer", 70),
		(u"Received Other", 70),
		(u"Studio Out", 50),
		(u"Placed in IQ", 50),
		(u"Notes", 200),
	]

	for col_num in xrange(len(columns)):
		c = ws.cell(row=row_num + 1, column=col_num + 1)
		c.value = columns[col_num][0]
		c.style.font.bold = False
		# set column width
		ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

	for obj in queryset:
		row_num += 1
		row = [
			obj.ad_date,
			obj.item_no,
			obj.mfg,
			obj.desc,
			obj.vendor_number,
			obj.order_date,
			obj.received_dc,
			obj.received_buyer,
			obj.received_other,
			obj.studio_out,
			obj.confirmed_placed,
			obj.notes,
		]
		for col_num in xrange(len(row)):
			c = ws.cell(row=row_num + 1, column=col_num + 1)
			c.value = row[col_num]
			c.style.alignment.wrap_text = True

	wb.save(response)
	return response

export_xlsx.short_description = u"Export XLSX"


def display_record(request, sku, item_no):
			user = request.user
			records = Product.objects.filter(item_no=item_no)
			printrecords = PrintFile.objects.filter(assoc_sku__iexact=sku)
			lgrecords = LargeWebfiles.objects.filter( Q(long_sku__icontains=item_no) | Q(sku__icontains=item_no))
			rgrecords = RegularWebfiles.objects.filter( Q(long_sku__icontains=item_no) | Q(sku__icontains=item_no) )
			msrecords = MSWebfiles.objects.filter( Q(long_sku__icontains=item_no) | Q(sku__icontains=item_no) )
			threcords = ThumbWebfiles.objects.filter( Q(long_sku__icontains=item_no) | Q(sku__icontains=item_no) )
	
			if request.user.is_authenticated():
					tpl = 'display_template.html'
					return render(request, tpl, {'user': user, 'records': records, 'lgrecords': lgrecords, 'rgrecords': rgrecords, 'msrecords': msrecords, 'threcords': threcords, 'printrecords': printrecords})
			else:
					tpl = 'display_template.html'
					return render(request, tpl, { 'records': records, 'printrecords': printrecords })

def file_upload(request):
				return render_to_response('upload.html')


def quicksearch(request, searchstring):
	if searchstring.find(' ') == -1:
		searchstring = searchstring.replace('_', '').replace('-', '')
		searchstring = searchstring[:4] + ' ' +  searchstring[4:9] + ' ' + searchstring[9:13]
	
		sku = searchstring[:4] + ' ' +  searchstring[5:10] + ' ' + searchstring[11:15]
	else:
		param = searchstring[:15]
		if not any(c.isalpha() for c in param):
			sku = param

	user = request.user
	# records = Product.objects.filter(item_no=item_no)
	printrecords = PrintFile.objects.filter(assoc_sku=sku)
	lgrecords = LargeWebfiles.objects.filter(sku__iexact=sku)
	rgrecords = RegularWebfiles.objects.filter(sku__iexact=sku)
	msrecords = MSWebfiles.objects.filter(sku__iexact=sku)
	threcords = ThumbWebfiles.objects.filter(sku__iexact=sku)
	   
	tpl = 'console_template.html'
	dict_ = {'user': user, 'lgrecords': lgrecords, 'rgrecords': rgrecords, 'msrecords': msrecords, 'threcords': threcords, 'printrecords': printrecords}
	return render(request, tpl, dict_)

def watch_list(request):
	user = request.user
	user_id = request.user.id
	records = HotItem.objects.all()
	tpl = 'mywatched.html'
	return render(request, tpl, {'user': user, 'records': records })


def display_watchlist(request, id):
	if request.method == 'POST':
		form = HotItemForm(request.POST)
		if form.is_valid():
			form.save()
			return HttpResponseRedirect('/mywatched/'+ id +'/')
		else:
			user = request.user
			profile = request.user.profile
			form = HotItemForm(initial={"watched_by": profile})
			return HttpResponseRedirect('/mywatched/'+ id +'/')
			
	else:
		user = request.user
		profile = request.user.profile
		form = HotItemForm(initial={"watched_by": profile})
		records = HotItem.objects.filter(watched_by_id=id)
		returndict = {'form': form, 'records': records, 'user': user}
		return render(request, 'mywatched.html', returndict)


def display_hotlist(request):
	if request.method == 'POST':
		form = HotItemForm(request.POST)
		if form.is_valid():
			form.save()
			return HttpResponseRedirect('../hotlist/')
		else:
			user = request.user
			profile = request.user.profile
			form = HotItemForm()
			return HttpResponseRedirect('/hotlist/')
			
	else:
		form = HotItemForm()
		records = HotItem.objects.all()
		returndict = {'form': form, 'records': records}
		return render(request, 'hotwatched.html', returndict)


def display_adminwatchlist(request, id):
	if request.method == 'POST':
		form = HotItemForm(request.POST)
		if form.is_valid():
			form.save()
			return HttpResponseRedirect('../adminwatched/')
		else:
			user = request.user
			profile = request.user.profile
			form = HotItemForm(initial={"watched_by": profile})
			return HttpResponseRedirect('/adminwatched/')
			
	else:
		user = request.user
		profile = request.user.profile
		form = HotItemForm(initial={"watched_by": profile})
		records = HotItem.objects.all()
		returndict = {'form': form, 'records': records, 'user': user}
		return render(request, 'adminwatched.html', returndict)


def delete_watchlist_item(request, id, userid):
	if request.method == 'POST':
		a=HotItem.objects.get(pk=id)
		form =  HotItemForm(request.POST, instance=a)
		if request.POST.get('delete'):
			a.delete()
			a=HotItem.objects.get(pk=id)
			form =  HotItemForm(instance=a)
			user = request.user
			user_id = request.user.id
			records = HotItem.objects.filter(watched_by_id=id)
			returndict = {'form': form, 'records': records, 'user': user}
			return render(request, 'mywatched.html', returndict)
		if form.is_valid():
			form.save()
			return HttpResponseRedirect('/mywatched/'+ id +'/')
	else:
		a=HotItem.objects.get(pk=id)
		form =  HotItemForm(instance=a)
		user = request.user
		user_id = request.user.id
		records = HotItem.objects.filter(watched_by_id=id)
		returndict = {'form': form, 'records': records, 'user': user}
		return render(request, 'mywatched.html', returndict)


def delete_adminwatchlist_item(request, id, userid):
	if request.method == 'POST':
		a=HotItem.objects.get(pk=id)
		form =  HotItemForm(request.POST, instance=a)
		if request.POST.get('delete'):
			a.delete()
			a=HotItem.objects.get(pk=id)
			form =  HotItemForm(instance=a)
			user = request.user
			user_id = request.user.id
			records = HotItem.objects.all()
			returndict = {'form': form, 'records': records, 'user': user}
			return render(request, 'adminwatched.html', returndict)
		if form.is_valid():
			form.save()
			return HttpResponseRedirect('/adminwatched/'+ id +'/')
	else:
		a=HotItem.objects.get(pk=id)
		form =  HotItemForm(instance=a)
		user = request.user
		user_id = request.user.id
		records = HotItem.objects.all()
		returndict = {'form': form, 'records': records, 'user': user}
		return render(request, 'adminwatched.html', returndict)


def delete_hotlist_item(request, id):
	if request.method == 'POST':
		a=HotItem.objects.get(pk=id)
		form =  HotItemForm(request.POST, instance=a)
		if request.POST.get('delete'):
			a.delete()
			form =  HotItemForm()
			records = HotItem.objects.all()
			returndict = {'form': form, 'records': records}
			return HttpResponseRedirect('/hotlist/')
		if form.is_valid():
			form.save()
			return HttpResponseRedirect('/hotlist/')
		else:
			return HttpResponseRedirect('/main/')
	else:
		a=HotItem.objects.get(pk=id)
		form =  HotItemForm(instance=a)
		records = HotItem.objects.all()
		returndict = {'form': form, 'records': records}
		return render(request, 'hotwatched.html', returndict)
		
def docs_adsheet(request):
	user = request.user
	tpl = 'docs/docs_adsheet.html'
	return render(request, tpl, {'user': user })

def docs_daily(request):
	user = request.user
	tpl = 'docs/docs_daily.html'
	return render(request, tpl, {'user': user })
	
def docs_deploy(request):
	user = request.user
	tpl = 'docs/docs_deploy.html'
	return render(request, tpl, {'user': user })

def docs_importer(request):
	user = request.user
	tpl = 'docs/docs_importer.html'
	return render(request, tpl, {'user': user })

def docs_inventory(request):
	user = request.user
	tpl = 'docs/docs_inventory.html'
	return render(request, tpl, {'user': user })

def docs_restore(request):
	user = request.user
	tpl = 'docs/docs_restore.html'
	return render(request, tpl, {'user': user })

def docs_system(request):
	user = request.user
	tpl = 'docs/itrac_doc.html'
	return render(request, tpl, {'user': user })

def studio_check(request):
	records = Product.objects.all()
	warnrecords = []
	for record in records:
		if record.studio_out is None:
			continue
		elif all([record.confirmed_placed is None, record.studio_out <= (date - timedelta(days=14)), record.studio_out >= date - timedelta(days=160)]):
			if len(record.item_no) < 16:
				for item in records:
					if item.sku == record.item_no:
						warnrecords.append(item)
			else:
				warnrecords.append(record)

	form = PostsearchForm()
	user = request.user
	records = warnrecords
	watchrecords = HotItem.objects.all()
	tmrecords = InboxEntry.objects.filter( Q(assigned_to=user.id) & Q(priority__iexact='urgent') )
	tpl = "studio_check.html"
	variables = {'user': user, 'form': form, 'watchrecords': watchrecords, 'tmrecords': tmrecords, 'records': records }
	return render(request, tpl, variables)

def is_consecutive(last_num,current_num):
	if current_num - last_num == 1:
		return True



def check_sequence(request):
	form = PostsearchForm()
	user = request.user 
	availableproductnums = []
	availablemodelnums = []
	availablealtnums = []
	imagelinks = []
	show_results = True
	if 'query' in request.GET:
			show_results = True
			query = request.GET['query'].strip()
			form = PostsearchForm({'query' : query})
			imagelinks = LargeWebfiles.objects.filter(Q(long_sku__icontains=query) | Q(sku__icontains=query))


	if    len(imagelinks) >= 1:
			sequence = []
			for image in imagelinks:
				sq_num = image.filename
				sequence.append(int(sq_num[-6:-4]))
			sequence = sorted(sequence)

			records = sequence
			default_dict = {'user': user, 'records': records, 'form': form}
			return render(request, 'sequence_result.html', default_dict)

	else:               
			tpl = "sequence_result.html"
			variables =  {'user': user, 'form': form, 'show_results': show_results}
			return render(request, tpl, variables)

def buyer_class(request):
	records = Buyers.objects.annotate(num_products=Count('product')) 
	tmp = 'buyerclass.html'
	variables = {'records': records }
	return render(request, tmp, variables)
